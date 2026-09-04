import openpyxl, sys, json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# Usage: python fill_excel.py template.xlsm out.xlsx '{"date":"04-09-2026","invoiceNo":"9999","client":"Test","description":"Desc","qty":"1","rate":"100","brand":"B","words":"One Only"}'
tpl = sys.argv[1]
out = sys.argv[2]
data = json.loads(sys.argv[3])
wb = openpyxl.load_workbook(tpl, keep_vba=False)
# Keep only Sales Invoice
keep='Sales Invoice'
for name in list(wb.sheetnames):
    if name!=keep:
        ws=wb[name]
        wb.remove(ws)
ws=wb[keep]
# Remove images (2 logos on right) - clear _images
try:
    ws._images = []
except:
    pass
# Fix AutoFilter/Table - remove to avoid Repaired Records error
try:
    ws.auto_filter = None
except:
    pass
# Clear tables correctly (openpyxl _tables is dict-like)
try:
    if hasattr(ws, '_tables'):
        try:
            ws._tables.clear()
        except:
            ws._tables = {}
except:
    pass
try:
    if hasattr(ws, 'tables'):
        try:
            ws.tables.clear()
        except:
            pass
except:
    pass
# Hide extra rows 20-38
try:
    for r in range(20, 39):
        ws.row_dimensions[r].hidden = True
        ws.row_dimensions[r].height = 0
    ws.print_area = 'A1:H44'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
except:
    pass
# Re-apply header style for B18:H18 (lost when tables removed) - dark header white text
try:
    hdr_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    hdr_font = Font(name="Century Gothic", size=11, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="E2E8F0")
    hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in ['B','C','D','E','F','G','H']:
        c = ws[col+'18']
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align
        c.border = hdr_border
    ws.row_dimensions[18].height = 18
except:
    pass
# Fill header
try: ws['H6'] = data.get('date','')
except: pass
try: ws['H7'] = str(data.get('invoiceNo',''))
except: pass
try: ws['F10'] = data.get('client','Walk-in Client')
except: pass
# Items - multi
items = data.get('items')
if not items:
    # fallback single
    items = [{'brand': data.get('brand',''), 'description': data.get('description',''), 'qty': data.get('qty',0), 'rate': data.get('rate',0), 'disc': 0}]
# Normalize
norm=[]
for it in items:
    try:
        qty = float(it.get('qty',0) or 0)
        rate = float(it.get('rate',0) or 0)
        disc = float(it.get('disc',0) or 0)
        line = qty*rate - disc
        norm.append({'brand': it.get('brand',''), 'description': it.get('description',''), 'qty': qty, 'rate': rate, 'disc': disc, 'line': line})
    except:
        pass
if not norm:
    norm=[{'brand':'','description':'','qty':0,'rate':0,'disc':0,'line':0}]
# Fill rows B19:H.. with alternating blue/white 12pt
blue_fill = PatternFill(start_color="D5E0EA", end_color="D5E0EA", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
light_border = Border(left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"), top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"))
bottom_thick = Border(left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"), top=Side(style="thin", color="E2E8F0"), bottom=Side(style="medium", color="0F172A"))
for idx, it in enumerate(norm):
    r = 19 + idx
    fill = white_fill if (idx % 2 == 0) else blue_fill
    is_last = (idx == len(norm)-1)
    border = bottom_thick if is_last else light_border
    try: ws[f'B{r}'] = idx+1
    except: pass
    try: ws[f'C{r}'] = it['brand']
    except: pass
    try: ws[f'D{r}'] = it['description']
    except: pass
    try: ws[f'E{r}'] = it['qty']
    except: pass
    try: ws[f'F{r}'] = it['rate']
    except: pass
    try: ws[f'G{r}'] = it['disc']
    except: pass
    try: ws[f'H{r}'] = it['line']
    except: pass
    ws.row_dimensions[r].hidden=False
    ws.row_dimensions[r].height=16
    for col in ['B','C','D','E','F','G','H']:
        c=ws[f'{col}{r}']
        c.fill=fill
        c.font=Font(name="Century Gothic", size=12, color="0F172A")
        c.alignment=Alignment(horizontal="center" if col in ['B','E','F','G','H'] else "left", vertical="center")
        c.border=border
# Hide remaining rows after items
for r in range(19+len(norm), 39):
    try:
        ws.row_dimensions[r].hidden=True
        ws.row_dimensions[r].height=0
        for col in ['B','C','D','E','F','G','H']:
            ws[f'{col}{r}'].value=None
    except: pass
subtotal = sum(it['line'] for it in norm)
try: ws['H41'] = subtotal
except: pass
try: ws['H43'] = subtotal
except: pass
try: ws['D44'] = data.get('words','')
except: pass
# H41/H43 blue like subtotal
try:
    for c in [ws['H41'], ws['H43']]:
        c.fill=blue_fill
        c.font=Font(name="Century Gothic", size=11, bold=True, color="0F172A")
except: pass
# Ensure no VBA for xlsx
try:
    wb.vba_archive = None
except:
    pass
wb.save(out)
print('saved')
